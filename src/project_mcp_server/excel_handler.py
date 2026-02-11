"""
Excel操作モジュール

openpyxlを使用してExcelファイルの読み書きを行います。
"""
import logging
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# カスタム例外クラス
class ExcelHandlerError(Exception):
    """Excel操作の基底例外クラス"""
    pass


class FileNotFoundError(ExcelHandlerError):
    """ファイルが見つからない"""
    pass


class FileAccessError(ExcelHandlerError):
    """ファイルにアクセスできない"""
    pass


class InvalidDataError(ExcelHandlerError):
    """データが不正"""
    pass


class ProjectNotFoundError(ExcelHandlerError):
    """プロジェクトが見つからない"""
    pass


class DuplicateProjectError(ExcelHandlerError):
    """プロジェクト名が重複"""
    pass


class ExcelHandler:
    """Excelファイルの読み書きを管理するクラス"""
    
    # 許可された進行状況
    VALID_STATUSES = ["未着手", "進行中", "完了", "中止"]
    
    # デフォルトのカラム名
    DEFAULT_COLUMNS = ["プロジェクト", "進行状況", "期限", "担当者", "備考"]
    
    def __init__(self, file_path: str):
        """
        ExcelHandlerを初期化します。
        
        Args:
            file_path: Excelファイルのパス
        """
        self.file_path = Path(file_path)
        logging.info(f"ExcelHandler initialized with file: {self.file_path}")
    
    def _load_workbook(self) -> Workbook:
        """
        Excelファイルを読み込みます（内部メソッド）。
        
        Returns:
            openpyxl.Workbook オブジェクト
        
        Raises:
            FileNotFoundError: ファイルが見つからない場合
            FileAccessError: ファイルにアクセスできない場合
        """
        if not self.file_path.exists():
            raise FileNotFoundError(f"ファイルが見つかりません: {self.file_path}")
        
        try:
            wb = openpyxl.load_workbook(self.file_path)
            logging.debug(f"Workbook loaded: {self.file_path}")
            return wb
        except PermissionError:
            raise FileAccessError(f"ファイルにアクセスできません。他のプログラムで使用中の可能性があります: {self.file_path}")
        except Exception as e:
            raise FileAccessError(f"ファイルの読み込みに失敗しました: {e}")
    
    def _save_workbook(self, wb: Workbook):
        """
        Excelファイルを保存します（内部メソッド）。
        
        Args:
            wb: 保存するWorkbookオブジェクト
        
        Raises:
            FileAccessError: ファイルにアクセスできない場合
        """
        try:
            wb.save(self.file_path)
            logging.debug(f"Workbook saved: {self.file_path}")
        except PermissionError:
            raise FileAccessError(f"ファイルを保存できません。他のプログラムで使用中の可能性があります: {self.file_path}")
        except Exception as e:
            raise FileAccessError(f"ファイルの保存に失敗しました: {e}")
    
    def _get_header_mapping(self, ws: Worksheet) -> Dict[str, int]:
        """
        ヘッダー行からカラム名とインデックスのマッピングを取得します。
        
        Args:
            ws: Worksheetオブジェクト
        
        Returns:
            カラム名をキー、列インデックス（0始まり）を値とする辞書
        """
        headers = {}
        for idx, cell in enumerate(ws[1]):
            if cell.value:
                headers[cell.value] = idx
        logging.debug(f"Header mapping: {headers}")
        return headers
    
    def _validate_status(self, status: str) -> bool:
        """
        進行状況の妥当性を検証します。
        
        Args:
            status: 検証する進行状況
        
        Returns:
            有効な場合True、無効な場合False
        """
        return status in self.VALID_STATUSES
    
    def get_all_projects(self) -> List[Dict]:
        """
        全プロジェクトを取得します。
        
        Returns:
            プロジェクト情報の配列
        
        Raises:
            FileNotFoundError: ファイルが見つからない場合
            FileAccessError: ファイルにアクセスできない場合
        """
        wb = self._load_workbook()
        ws = wb.active
        
        # ヘッダー行を取得
        headers = [cell.value for cell in ws[1]]
        
        # データ行を辞書化
        projects = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 空行をスキップ
            if not any(row):
                continue
            
            project = {}
            for idx, header in enumerate(headers):
                if header and idx < len(row):
                    project[header] = row[idx]
            
            projects.append(project)
        
        logging.info(f"{len(projects)}件のプロジェクトを取得しました")
        return projects
    
    def get_project(self, project_name: str) -> Optional[Dict]:
        """
        特定のプロジェクトを取得します。
        
        Args:
            project_name: プロジェクト名
        
        Returns:
            プロジェクト情報の辞書、見つからない場合はNone
        """
        projects = self.get_all_projects()
        
        for project in projects:
            if project.get("プロジェクト") == project_name:
                logging.info(f"プロジェクトを取得しました: {project_name}")
                return project
        
        logging.warning(f"プロジェクトが見つかりません: {project_name}")
        return None
    
    def add_project(self, project_data: Dict) -> bool:
        """
        新しいプロジェクトを追加します。
        
        Args:
            project_data: プロジェクト情報の辞書
        
        Returns:
            成功した場合True
        
        Raises:
            InvalidDataError: データが不正な場合
            DuplicateProjectError: 同名のプロジェクトが存在する場合
            FileAccessError: ファイルにアクセスできない場合
        """
        # バリデーション
        project_name = project_data.get("プロジェクト")
        if not project_name:
            raise InvalidDataError("プロジェクト名は必須です")
        
        status = project_data.get("進行状況", "未着手")
        if not self._validate_status(status):
            raise InvalidDataError(
                f"進行状況は {', '.join(self.VALID_STATUSES)} のいずれかを指定してください"
            )
        
        # 重複チェック
        existing = self.get_project(project_name)
        if existing:
            raise DuplicateProjectError(f"同名のプロジェクトが既に存在します: {project_name}")
        
        # Excel追加
        wb = self._load_workbook()
        ws = wb.active
        
        # ヘッダーマッピングを取得
        header_mapping = self._get_header_mapping(ws)
        
        # 新規行のデータを準備
        new_row = [""] * len(header_mapping)
        for column_name, column_idx in header_mapping.items():
            if column_name in project_data:
                new_row[column_idx] = project_data[column_name]
        
        # 行を追加
        ws.append(new_row)
        self._save_workbook(wb)
        
        logging.info(f"プロジェクトを追加しました: {project_name}")
        return True
    
    def update_project(self, project_name: str, updates: Dict) -> bool:
        """
        既存プロジェクトを更新します。
        
        Args:
            project_name: 更新対象のプロジェクト名
            updates: 更新する内容の辞書
        
        Returns:
            成功した場合True
        
        Raises:
            ProjectNotFoundError: プロジェクトが見つからない場合
            InvalidDataError: データが不正な場合
            FileAccessError: ファイルにアクセスできない場合
        """
        # 進行状況のバリデーション
        if "進行状況" in updates:
            if not self._validate_status(updates["進行状況"]):
                raise InvalidDataError(
                    f"進行状況は {', '.join(self.VALID_STATUSES)} のいずれかを指定してください"
                )
        
        wb = self._load_workbook()
        ws = wb.active
        
        # ヘッダーマッピングを取得
        header_mapping = self._get_header_mapping(ws)
        
        # プロジェクト名の列インデックスを取得
        project_col_idx = header_mapping.get("プロジェクト")
        if project_col_idx is None:
            raise InvalidDataError("Excelファイルに「プロジェクト」列が見つかりません")
        
        # 対象行を検索
        target_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[project_col_idx].value == project_name:
                target_row = row_idx
                break
        
        if target_row is None:
            raise ProjectNotFoundError(f"プロジェクトが見つかりません: {project_name}")
        
        # 更新
        for column_name, new_value in updates.items():
            if column_name in header_mapping:
                column_idx = header_mapping[column_name]
                ws.cell(row=target_row, column=column_idx + 1, value=new_value)
        
        self._save_workbook(wb)
        
        logging.info(f"プロジェクトを更新しました: {project_name}")
        return True
    
    def delete_project(self, project_name: str) -> bool:
        """
        プロジェクトを削除します。
        
        Args:
            project_name: 削除するプロジェクト名
        
        Returns:
            成功した場合True
        
        Raises:
            ProjectNotFoundError: プロジェクトが見つからない場合
            FileAccessError: ファイルにアクセスできない場合
        """
        wb = self._load_workbook()
        ws = wb.active
        
        # ヘッダーマッピングを取得
        header_mapping = self._get_header_mapping(ws)
        
        # プロジェクト名の列インデックスを取得
        project_col_idx = header_mapping.get("プロジェクト")
        if project_col_idx is None:
            raise InvalidDataError("Excelファイルに「プロジェクト」列が見つかりません")
        
        # 対象行を検索
        target_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[project_col_idx].value == project_name:
                target_row = row_idx
                break
        
        if target_row is None:
            raise ProjectNotFoundError(f"プロジェクトが見つかりません: {project_name}")
        
        # 行を削除
        ws.delete_rows(target_row)
        self._save_workbook(wb)
        
        logging.info(f"プロジェクトを削除しました: {project_name}")
        return True
    
    def search_projects(self, filters: Dict) -> List[Dict]:
        """
        条件に合致するプロジェクトを検索します。
        
        Args:
            filters: 検索条件の辞書（例: {"進行状況": "進行中", "担当者": "田中"}）
        
        Returns:
            条件に合致するプロジェクトの配列
        """
        all_projects = self.get_all_projects()
        results = []
        
        for project in all_projects:
            match = True
            for key, value in filters.items():
                if project.get(key) != value:
                    match = False
                    break
            
            if match:
                results.append(project)
        
        logging.info(f"検索結果: {len(results)}件のプロジェクトが見つかりました")
        return results
